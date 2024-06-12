import json
import os
import shutil

import pandas as pd
from openpyxl import load_workbook

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

# def read_csv(filepath):
#     return pd.read_csv(filepath)

# def read_excel(filepath, sheet_name=0):
#     return pd.read_excel(filepath, sheet_name=sheet_name)

# def read_text(filepath, delimiter="\t"):
#     return pd.read_csv(filepath, delimiter=delimiter)

# def save_to_csv(df, filepath, index=True):
#     df.to_csv(filepath, index=index)

# def save_to_excel(df, filepath, sheet_name='Sheet1', index=True):
#     df.to_excel(filepath, sheet_name=sheet_name, index=index)


def get_sheetnames(fpath):
    """
    List all sheet names of an input Excel file (xlsx).

    Args:
    fpath (str): Path to the Excel file.

    Returns:
    list: A list of sheet names if the file exists and is an Excel file.
    None: If the file does not exist or is not an Excel file.
    """
    # Check if the file exists
    if not os.path.exists(fpath):
        print(f"The file {fpath} does not exist.")
        return None

    # Check if the file is an Excel file
    if not fpath.lower().endswith(".xlsx"):
        print(f"The file {fpath} is not an Excel file.")
        return None

    # Read the Excel file and get sheet names
    try:
        excel_file = pd.ExcelFile(fpath)
        sheet_names = excel_file.sheet_names
        return sheet_names
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return None


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


def save_df_to_excel(
    df_to_save,
    filepath,
    sheet_name,
    mode="a",
    if_sheet_exists="replace",
    index=False,
    verbose=True,
):
    """
    Save a pandas DataFrame to a specified sheet in an Excel file.
    This function can create a new file, append a new sheet, or replace an existing sheet.

    Args:
    df_to_save (pd.DataFrame): The pandas DataFrame to be saved.
    filepath (str): The file path where the Excel file is located or will be created.
    sheet_name (str): The name of the sheet where the DataFrame will be saved.
    mode (str): The mode for opening the Excel file. 'a' for append, 'w' for write. Default is 'a'.
    if_sheet_exists (str): The behavior when the sheet already exists. Options are 'replace', 'new', 'skip'. Default is 'replace'.
    index (bool): Whether to include DataFrame index. Default is False.
    verbose (bool): Whether to print success message. Default is True.

    Returns:
    None: The function writes the DataFrame to the Excel file and does not return anything.
    """
    try:
        # Check if the file exists
        file_exists = os.path.isfile(filepath)

        # If file doesn't exist, create a new one
        if not file_exists:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df_to_save.to_excel(writer, sheet_name=sheet_name, index=index)
            if verbose:
                print(
                    f"Excel file created and DataFrame written to '{sheet_name}' at {filepath}."
                )
            return

        # If file exists, load it and append the new sheet
        if file_exists and mode == "a":
            book = load_workbook(filepath)
            with pd.ExcelWriter(
                filepath, engine="openpyxl", mode="a"
            ) as writer:
                writer.book = book

                # Check if sheet already exists
                if sheet_name in writer.book.sheetnames:
                    if if_sheet_exists == "replace":
                        idx = writer.book.sheetnames.index(sheet_name)
                        std = writer.book.get_sheet_by_name(sheet_name)
                        writer.book.remove(std)
                        writer.book.create_sheet(sheet_name, idx)
                    elif if_sheet_exists == "skip":
                        if verbose:
                            print(
                                f"Sheet '{sheet_name}' already exists and was skipped."
                            )
                        return
                    elif if_sheet_exists == "new":
                        sheet_name = sheet_name + "_new"

                df_to_save.to_excel(writer, sheet_name=sheet_name, index=index)
                if verbose:
                    print(
                        f"DataFrame successfully written to '{sheet_name}' in the Excel file at {filepath}."
                    )

        # If file exists and mode is 'w', create a new file
        if file_exists and mode == "w":
            with pd.ExcelWriter(
                filepath, engine="openpyxl", mode="w"
            ) as writer:
                df_to_save.to_excel(writer, sheet_name=sheet_name, index=index)
            if verbose:
                print(
                    f"Excel file created and DataFrame written to '{sheet_name}' at {filepath}."
                )

    except Exception as e:
        print(f"An error occurred while writing to the Excel file: {e}")


def save_dict_to_json(
    dictionary, folder_path, file_name, indent=4, sort_keys=True, overwrite=True
):
    """
    Save a Python dictionary to a JSON file.

    Args:
    dictionary (dict): The Python dictionary to be saved.
    folder_path (str): The path to the folder where the JSON file will be saved.
    file_name (str): The name of the JSON file (without extension).
    indent (int, optional): The indentation level of the JSON file. Defaults to 4.
    sort_keys (bool, optional): Whether to sort the keys in the JSON file. Defaults to True.
    overwrite (bool, optional): Whether to overwrite an existing file with the same name. Defaults to True.

    Returns:
    str: The path to the saved JSON file.
    """
    if not isinstance(dictionary, dict):
        raise ValueError("The 'dictionary' parameter must be a dictionary.")

    if not os.path.isdir(folder_path):
        os.makedirs(folder_path)

    file_path = os.path.join(folder_path, f"{file_name}.json")

    if not overwrite and os.path.exists(file_path):
        raise FileExistsError(
            f"The file '{file_path}' already exists and overwrite is set to False."
        )

    with open(file_path, "w") as json_file:
        json.dump(dictionary, json_file, indent=indent, sort_keys=sort_keys)

    return file_path


def read_json_to_dict(file_path):
    """
    Read a JSON file and convert its contents to a Python dictionary.

    Args:
    file_path (str): The path to the JSON file to be read.

    Returns:
    dict: The contents of the JSON file as a Python dictionary.

    Raises:
    FileNotFoundError: If the file does not exist.
    ValueError: If the file contents are not valid JSON.
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")

    with open(file_path, "r") as json_file:
        try:
            dictionary = json.load(json_file)
        except json.JSONDecodeError as e:
            raise ValueError(
                f"Error decoding JSON from the file '{file_path}': {e}"
            )

    if not isinstance(dictionary, dict):
        raise ValueError(
            f"The file '{file_path}' does not contain a valid dictionary."
        )

    return dictionary

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def save_configs(src_file, dest_folder, verbose=True):
    """
    Save a file to a destination folder, overwriting the existing file if the source file is newer.

    Parameters:
    src_file (str): The path to the source file to be saved.
    dest_folder (str): The path to the destination folder where the file should be saved.
    verbose (bool): Optional; If True, prints information about the operation (default is True).

    Returns:
    str: A message indicating the action taken (e.g., "File saved", "File overwritten", "No action taken").

    Raises:
    FileNotFoundError: If the source file does not exist.
    """
    if not os.path.exists(src_file):
        raise FileNotFoundError(f"Source file '{src_file}' not found.")

    # Ensure the destination folder exists
    os.makedirs(dest_folder, exist_ok=True)

    # Construct the destination file path
    dest_file = os.path.join(dest_folder, os.path.basename(src_file))

    action_message = ""

    # Check if the destination file exists
    if os.path.exists(dest_file):
        # Only copy if the source file is newer
        if os.path.getmtime(src_file) > os.path.getmtime(dest_file):
            shutil.copy2(src_file, dest_file)
            action_message = f"Overwriting existing file '{dest_file}' with newer '{src_file}'."
        else:
            action_message = f"Existing file '{dest_file}' is up-to-date. No action taken."
    else:
        # If the file does not exist, copy it to the destination folder
        shutil.copy2(src_file, dest_file)
        action_message = f"Saving new file '{src_file}' to '{dest_folder}'."

    # Print the action message if verbose is True
    if verbose:
        print(action_message)

    return action_message